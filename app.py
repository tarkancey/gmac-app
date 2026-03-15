import streamlit as st
import requests
from scipy.stats import poisson
import pandas as pd
from datetime import datetime, timedelta, timezone
import io
from openpyxl.styles import PatternFill

# --- SAYFA AYARLARI ---
st.set_page_config(page_title="GMAC V11.0 - Renk Kodlu Value Sistemi", page_icon="⚔️", layout="wide")

if "analiz_df" not in st.session_state:
    st.session_state.analiz_df = None

# Tüm hedef ligler (Alt ligler ve Kupalar dahil)
TARGET_IDS = [203, 204, 39, 40, 140, 141, 78, 79, 135, 136, 61, 62, 88, 89, 94, 144, 119, 120, 121, 179, 345, 197, 106, 210, 211, 212, 2, 3, 218, 207, 848]

def fix_timezone(date_str):
    try:
        if date_str.endswith('Z'): date_str = date_str.replace('Z', '+00:00')
        dt_obj = datetime.fromisoformat(date_str)
        tr_zone = timezone(timedelta(hours=3))
        dt_tr = dt_obj.astimezone(tr_zone)
        return dt_tr.strftime("%Y-%m-%d"), dt_tr.strftime("%H:%M") 
    except: return date_str[:10], date_str[11:16]

@st.cache_data(ttl=3600)
def get_league_standings(lig_id, season, api_key):
    headers = {"x-apisports-key": api_key}
    try:
        resp = requests.get("https://v3.football.api-sports.io/standings", headers=headers, params={"league": lig_id, "season": season})
        return {t['team']['id']: t['points'] for g in resp.json().get('response', [])[0]['league']['standings'] for t in g} if resp.json().get('response') else {}
    except: return {}

def get_matches_range(lig_id, season_year, start_date, end_date, api_key):
    headers = {"x-apisports-key": api_key}
    try:
        return requests.get("https://v3.football.api-sports.io/fixtures", headers=headers, params={"league": lig_id, "season": season_year, "from": start_date, "to": end_date, "timezone": "Europe/Istanbul"}).json().get('response', [])
    except: return []

def get_odds(fixture_id, api_key):
    headers = {"x-apisports-key": api_key}
    url = "https://v3.football.api-sports.io/odds"
    odds_pool = {"MS1": [], "MSX": [], "MS2": [], "2.5U": [], "2.5A": [], "3.5U": [], "3.5A": [], "KGV": []}
    try:
        data = requests.get(url, headers=headers, params={"fixture": fixture_id}).json().get('response', [])
        if data:
            for bk in data[0].get('bookmakers', []):
                for bet in bk.get('bets', []):
                    if bet['id'] == 1: 
                        for v in bet['values']:
                            if v['value'] == "Home": odds_pool["MS1"].append(float(v['odd']))
                            elif v['value'] == "Draw": odds_pool["MSX"].append(float(v['odd']))
                            elif v['value'] == "Away": odds_pool["MS2"].append(float(v['odd']))
                    elif bet['id'] == 5: 
                        for v in bet['values']:
                            if v['value'] == "Over 2.5": odds_pool["2.5U"].append(float(v['odd']))
                            elif v['value'] == "Under 2.5": odds_pool["2.5A"].append(float(v['odd']))
                            elif v['value'] == "Over 3.5": odds_pool["3.5U"].append(float(v['odd']))
                            elif v['value'] == "Under 3.5": odds_pool["3.5A"].append(float(v['odd']))
                    elif bet['id'] == 8: 
                         for v in bet['values']:
                            if v['value'] == "Yes": odds_pool["KGV"].append(float(v['odd']))
    except: pass
    
    return {k: round(sum(v)/len(v), 2) if v else 0.0 for k, v in odds_pool.items()}

def get_stats(lig_id, team_id, season_year, api_key):
    headers = {"x-apisports-key": api_key}
    try:
        s = requests.get("https://v3.football.api-sports.io/teams/statistics", headers=headers, params={"league": lig_id, "team": team_id, "season": season_year}).json().get('response')
        return {"form": s.get('form', "")[-5:], "hf": float(s['goals']['for']['average']['home']), "ha": float(s['goals']['against']['average']['home']), "af": float(s['goals']['for']['average']['away']), "aa": float(s['goals']['against']['average']['away'])} if s else None
    except: return None

def get_h2h(ev_id, dep_id, api_key):
    headers = {"x-apisports-key": api_key}
    try:
        resp = requests.get("https://v3.football.api-sports.io/fixtures/headtohead", headers=headers, params={"h2h": f"{ev_id}-{dep_id}", "last": 5}).json().get('response', [])
        w, d, l = 0, 0, 0
        for match in resp:
            if match['fixture']['status']['short'] in ['FT', 'AET', 'PEN']:
                h_goals = match['goals']['home']
                a_goals = match['goals']['away']
                if h_goals == a_goals:
                    d += 1
                elif (match['teams']['home']['id'] == ev_id and h_goals > a_goals) or (match['teams']['away']['id'] == ev_id and a_goals > h_goals):
                    w += 1
                else:
                    l += 1
        return f"{w}-{d}-{l}"
    except:
        return "0-0-0"

def get_injuries(fixture_id, ev_id, dep_id, api_key):
    headers = {"x-apisports-key": api_key}
    try:
        resp = requests.get("https://v3.football.api-sports.io/injuries", headers=headers, params={"fixture": fixture_id}).json().get('response', [])
        ev_eksik = sum(1 for p in resp if p['team']['id'] == ev_id)
        dep_eksik = sum(1 for p in resp if p['team']['id'] == dep_id)
        return ev_eksik, dep_eksik
    except:
        return 0, 0

def calculate_momentum_xg(h_stats, a_stats, h_pts, a_pts):
    def form_multiplier(form_str):
        pts = (form_str.count('W') * 3) + (form_str.count('D') * 1)
        return 0.7 + (pts / 15.0) * 0.6 
    
    h_mom, a_mom = form_multiplier(h_stats['form']), form_multiplier(a_stats['form'])
    
    diff = h_pts - a_pts
    h_pts_multiplier = 1.0 + max(min(diff * 0.01, 0.3), -0.3)
    a_pts_multiplier = 1.0 + max(min(-diff * 0.01, 0.3), -0.3)
    
    ev_xg = ((h_stats['hf'] + a_stats['aa']) / 2.0) * h_mom * h_pts_multiplier
    dep_xg = ((h_stats['ha'] + a_stats['af']) / 2.0) * a_mom * a_pts_multiplier
    
    return max(0.1, ev_xg), max(0.1, dep_xg)

def calculate_hybrid_probabilities(ev_xg, dep_xg):
    ms1, msx, ms2, kg_var = 0, 0, 0, 0
    total_prob = 0
    for h in range(10):
        for a in range(10):
            p = poisson.pmf(h, ev_xg) * poisson.pmf(a, dep_xg)
            if h == a and h in [0, 1]: p *= 1.15
            total_prob += p
            if h > a: ms1 += p
            elif h == a: msx += p
            else: ms2 += p
            if h > 0 and a > 0: kg_var += p
            
    norm = 100.0 / total_prob if total_prob > 0 else 0
    total_xg = ev_xg + dep_xg
    prob_under_25 = poisson.cdf(2, total_xg) * 100
    prob_under_35 = poisson.cdf(3, total_xg) * 100
    if total_xg > 2.2: prob_under_35 *= 0.85 
    
    return {
        "1": ms1 * norm, "X": msx * norm, "2": ms2 * norm, 
        "KGV": kg_var * norm, 
        "2.5A": prob_under_25, "2.5U": 100 - prob_under_25, 
        "3.5A": prob_under_35, "3.5U": 100 - prob_under_35
    }

def calc_value(prob, odd):
    # DÜZELTİLDİ: Fonksiyon adı calc_value
    if odd == 0.0 or prob == 0.0: return 0.0
    return round(((prob / 100.0) * odd) - 1, 2)

# --- EKRAN RENKLENDİRME ---
def color_value(val):
    if isinstance(val, (int, float)):
        if val >= 0.05:
            return 'background-color: #c6efce; color: #006100;' # Yeşil
        elif val <= -0.15:
            return 'background-color: #ffc7ce; color: #9c0006;' # Kırmızı
    return ''

# --- ARAYÜZ (UI) ---
st.title("⚔️ GMAC V11.0 - Renk Kodlu Value Analizi")

with st.sidebar:
    st.header("⚙️ Ayarlar")
    api_key = st.text_input("API Key Giriniz:", type="password")
    
    st.header("📅 Tarih Seçimi")
    now = datetime.now()
    bugun_str = now.strftime("%Y-%m-%d")
    yarin_str = (now + timedelta(days=1)).strftime("%Y-%m-%d")
    
    secim = st.radio("Taramak İstediğiniz Gün:", ["Bugün", "Yarın"])
    hedef_tarih = bugun_str if secim == "Bugün" else yarin_str
    
    baslat = st.button("🚀 Analizi Başlat", type="primary")

if baslat:
    if not api_key:
        st.error("Lütfen sol menüden API Key giriniz!")
    else:
        all_excel_data = []
        headers = {"x-apisports-key": api_key}
        
        with st.status("Maçlar taranıyor, Value oranları hesaplanıyor...", expanded=True) as status:
            try:
                resp = requests.get("https://v3.football.api-sports.io/leagues", headers=headers, params={"current": "true"})
                valid_leagues = [{"id": i['league']['id'], "name": i['league']['name'], "year": i['seasons'][0]['year']} for i in resp.json().get('response', []) if i['league']['id'] in TARGET_IDS]
                
                progress_bar = st.progress(0)
                total_leagues = len(valid_leagues)
                
                for i, l in enumerate(valid_leagues):
                    st.write(f"⏳ Taranıyor: {l['name']}")
                    maclar = get_matches_range(l['id'], l['year'], hedef_tarih, hedef_tarih, api_key)
                    
                    for mac in maclar:
                        try:
                            fix_id = mac['fixture']['id']
                            tr_tarih, saat = fix_timezone(mac['fixture']['date'])
                            ev_ad, dep_ad = mac['teams']['home']['name'], mac['teams']['away']['name']
                            ev_id, dep_id, lig_id, sezon = mac['teams']['home']['id'], mac['teams']['away']['id'], mac['league']['id'], mac['league']['season']
                            
                            points_map = get_league_standings(lig_id, sezon, api_key)
                            ev_puan, dep_puan = points_map.get(ev_id, 0), points_map.get(dep_id, 0)
                            
                            ms_durumu = mac['fixture']['status']['short']
                            skor = f"{mac['goals']['home']}-{mac['goals']['away']}" if ms_durumu in ['FT', 'AET', 'PEN'] else ("NS" if ms_durumu == "NS" else ms_durumu)

                            h_s = get_stats(lig_id, ev_id, sezon, api_key)
                            a_s = get_stats(lig_id, dep_id, sezon, api_key)
                            
                            if h_s and a_s:
                                ev_eksik, dep_eksik = get_injuries(fix_id, ev_id, dep_id, api_key)
                                h2h_str = get_h2h(ev_id, dep_id, api_key)
                                
                                ev_xg, dep_xg = calculate_momentum_xg(h_s, a_s, ev_puan, dep_puan)
                                probs = calculate_hybrid_probabilities(ev_xg, dep_xg)
                                odds = get_odds(fix_id, api_key)
                                
                                tr_tarih_gosterim = datetime.strptime(tr_tarih, "%Y-%m-%d").strftime("%d.%m.%Y")
                                
                                # DÜZELTİLDİ: Fonksiyon adları burada calc_value olarak değiştirildi.
                                all_excel_data.append({
                                    "Tarih": tr_tarih_gosterim, "Sort_Tarih": tr_tarih, "Saat": saat, "Lig": mac['league']['name'],
                                    "Ev": ev_ad, "Dep": dep_ad, "Skor": skor, 
                                    "Ev Eksik": ev_eksik, "Dep Eksik": dep_eksik, 
                                    "Ev xG": round(ev_xg, 2), "Dep xG": round(dep_xg, 2),
                                    "Ev Form": h_s['form'], "Dep Form": a_s['form'], "H2H W-D-L": h2h_str,
                                    
                                    "MS1 Oran": odds["MS1"], "MS1 %": round(probs['1']), "MS1 VAL": calc_value(probs['1'], odds["MS1"]),
                                    "MSX Oran": odds["MSX"], "MSX %": round(probs['X']), "MSX VAL": calc_value(probs['X'], odds["MSX"]),
                                    "MS2 Oran": odds["MS2"], "MS2 %": round(probs['2']), "MS2 VAL": calc_value(probs['2'], odds["MS2"]),
                                    
                                    "KG Var Oran": odds["KGV"], "KG Var %": round(probs['KGV']), "KGV VAL": calc_value(probs['KGV'], odds["KGV"]),
                                    
                                    "2.5Ü Oran": odds["2.5U"], "2.5Ü %": round(probs['2.5U']), "2.5Ü VAL": calc_value(probs['2.5U'], odds["2.5U"]),
                                    "2.5A Oran": odds["2.5A"], "2.5A %": round(probs['2.5A']), "2.5A VAL": calc_value(probs['2.5A'], odds["2.5A"]),
                                    "3.5Ü Oran": odds["3.5U"], "3.5Ü %": round(probs['3.5U']), "3.5Ü VAL": calc_value(probs['3.5U'], odds["3.5U"]),
                                    "3.5A Oran": odds["3.5A"], "3.5A %": round(probs['3.5A']), "3.5A VAL": calc_value(probs['3.5A'], odds["3.5A"])
                                })
                        except Exception as e: 
                            pass
                    progress_bar.progress((i + 1) / total_leagues)
                status.update(label="Analiz Tamamlandı!", state="complete", expanded=False)
                
                if all_excel_data:
                    temp_df = pd.DataFrame(all_excel_data)
                    temp_df = temp_df.sort_values(by=["Sort_Tarih", "Saat", "Ev"], ascending=[True, True, True]).reset_index(drop=True)
                    temp_df = temp_df.drop(columns=["Sort_Tarih"])
                    st.session_state.analiz_df = temp_df
                else:
                    st.session_state.analiz_df = pd.DataFrame() 
                    
            except Exception as e:
                st.error(f"Bir hata oluştu: {e}")

if st.session_state.analiz_df is not None:
    df = st.session_state.analiz_df
    
    if not df.empty:
        st.success("✅ Analiz tamamlandı! Yeşille işaretlenmiş değerleri öncelikli olarak değerlendirebilirsiniz.")
        
        # Sadece VAL yazan sütunları renklendir (Streamlit ekranı için)
        val_columns = [col for col in df.columns if 'VAL' in col]
        styled_df = df.style.map(color_value, subset=val_columns)
        
        st.dataframe(styled_df, use_container_width=True, hide_index=True)
        
        # Excel Dosyasını Hazırlama ve Renklendirme
        buffer_all = io.BytesIO()
        with pd.ExcelWriter(buffer_all, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Analiz")
            worksheet = writer.sheets['Analiz']
            
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Hangi sütunların VAL olduğunu bul (Excel indexleri 1'den başlar)
            val_indices = [i + 1 for i, col in enumerate(df.columns) if 'VAL' in col]
            
            for row_idx, row in enumerate(df.itertuples(index=False), start=2): # 1. satır başlık
                for col_idx in val_indices:
                    cell_value = row[col_idx - 1]
                    if isinstance(cell_value, (int, float)):
                        if cell_value >= 0.05:
                            worksheet.cell(row=row_idx, column=col_idx).fill = green_fill
                        elif cell_value <= -0.15:
                            worksheet.cell(row=row_idx, column=col_idx).fill = red_fill

        st.download_button(
            label="📥 Renk Kodlu Tabloyu Excel Olarak İndir",
            data=buffer_all.getvalue(),
            file_name=f"GMAC_Value_Analizi_{datetime.now().strftime('%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
    else:
        st.warning("Seçilen tarihte taranacak maç bulunamadı.")
